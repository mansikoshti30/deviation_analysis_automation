import os
import geopandas as gpd 
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from shapely.ops import unary_union
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage


## reomve the duplicates from input data 
def get_data_preprocessed(act_exv, act_dump, pln_exv, pln_dump, output_dir):

    def clean_gdf(input_path, prefix, out_name, cleaned_folder):
        gdf = gpd.read_file(input_path)

        # Keep only geometry
        gdf = gdf[['geometry']]

        # Explode multipart geometries
        gdf = gdf.explode(index_parts=True)

        # Drop duplicates
        gdf = gdf.drop_duplicates(subset='geometry').reset_index(drop=True)

        # Add area in hectares
        gdf["AREA_in_ha"] = gdf.geometry.area / 10000.0

        # Add AREA_NAME
        gdf["AREA_NAME"] = [f"{prefix}_{i+1}" for i in range(len(gdf))]

        # Save to subfolder
        out_path = os.path.join(cleaned_folder, f"{out_name}.shp")
        gdf.to_file(out_path)

        return gdf

    # Create subfolder inside output_dir
    sub_folder = os.path.join(output_dir, "preprocessed_inputs")
    os.makedirs(sub_folder, exist_ok=True)

    ## actual data
    clean_act_exv = clean_gdf(act_exv, "actual_excavation_area", "actual_excavation", sub_folder)
    clean_act_dump = clean_gdf(act_dump, "actual_dump_area", "actual_dump", sub_folder)

    ## planned data
    clean_pln_exv = clean_gdf(pln_exv, "planned_excavation_area", "planned_excavation", sub_folder)
    clean_pln_dump = clean_gdf(pln_dump, "planned_dump_area", "planned_dump", sub_folder)


    # Extract EPSG from any input shapefile (e.g., act_exv)
    gdf_check = gpd.read_file(act_exv)
    epsg_code = gdf_check.crs.to_string() if gdf_check.crs else None

    return clean_act_exv, clean_act_dump, clean_pln_exv, clean_pln_dump, epsg_code


def split_planned_done_new(done_file, planned_file, output_dir, area_type):
    """
    - Saves 3 shapefiles: planned_and_done, unplanned_and_done, planned_and_not_done
    - Returns (results_dict, df) where df has ONE ROW PER FEATURE IN `done`:
        AREA_TYPE, AREA_NAME,
        Planned_and_Active (inHa)  -> area( done_i ∩ planned )
        Unplanned_and_Active (inHa)-> area( done_i - planned )
    """
    # Load GeoDataFrames
    done = done_file
    planned = planned_file

    # Ensure CRS match
    if done.crs != planned.crs:
        planned = planned.to_crs(done.crs)

    # ---- Create shapefiles (same as your workflow) ----
    planned_and_done = gpd.overlay(planned[["geometry"]], done[["geometry"]], how="intersection")
    unplanned_and_done = gpd.overlay(done[["geometry"]], planned[["geometry"]], how="difference")
    planned_and_not_done = gpd.overlay(planned[["geometry"]], done[["geometry"]], how="difference")

    results = {
        f"planned_and_done_{area_type}": planned_and_done,
        f"unplanned_and_done_{area_type}": unplanned_and_done,
        f"planned_and_not_done_{area_type}": planned_and_not_done
    }

    # Output folder - shape files
    shp_folder = os.path.join(output_dir, "shapefile_outputs")
    os.makedirs(shp_folder, exist_ok=True)

    # Output folder - geojson files
    geojson_folder = os.path.join(output_dir, "geojson_outputs")
    os.makedirs(geojson_folder, exist_ok=True)

    # Save each result as shapefile with clean names and attributes
    for name, gdf in results.items():
        if not gdf.empty:
            gdf = gdf[["geometry"]].reset_index(drop=True)
            gdf = gdf.explode(ignore_index=True)

            ## attributes
            gdf["AREA_in_HA"] = gdf.geometry.area / 10000.0
            gdf["AREA_NAME"] = [f"{name}_area_{i+1}" for i in range(len(gdf))]

            ## save as shapefile
            gdf.to_file(os.path.join(shp_folder, f"{name}.shp"))

            ## save as geojson
            gdf.to_file(os.path.join(geojson_folder,f"{name}.geojson"),driver="GeoJSON")


    # ---- Build summary df: one row per feature in done ----
    # Work with singlepart geometries of 'done'
    done_single = done[["geometry"]].explode(ignore_index=True).copy()

    # Optional fix invalid geoms: buffer(0)
    done_single["geometry"] = done_single.geometry.buffer(0)
    planned_union = unary_union(planned.geometry.buffer(0))

    summary_rows = []
    for i, geom in enumerate(done_single.geometry):
        if geom is None or geom.is_empty:
            # Skip empty geometry safely
            inter_area = 0.0
            done_area = 0.0
        else:
            # Intersection of THIS done feature with ALL planned
            inter_geom = geom.intersection(planned_union) if not getattr(planned_union, "is_empty", True) else None
            inter_area = 0.0 if (inter_geom is None or inter_geom.is_empty) else inter_geom.area
            done_area = geom.area

        # Areas in hectares
        inter_ha = inter_area / 10000.0
        unplanned_ha = max((done_area - inter_area), 0.0) / 10000.0



        summary_rows.append({
            "AREA_TYPE": area_type,
            "AREA_NAME": f"{area_type}_area_{i+1}",
            "Compliant_Area (Ha)": inter_ha,
            "Deviation_Area (Ha)": unplanned_ha
        })

    df = pd.DataFrame(summary_rows)

    return results, df


def merge_and_save(df1, df2, output_path):
    
    # Merge/append
    merged_df = pd.concat([df1, df2], ignore_index=True)

    # Decide output file name
    if os.path.isdir(output_path):
        output_file = os.path.join(output_path, "merged_summary.xlsx")
    else:
        # Ensure extension is .xlsx
        if not output_path.lower().endswith(".xlsx"):
            output_file = f"{output_path}.xlsx"
        else:
            output_file = output_path

    # Save to Excel
    merged_df.to_excel(output_file, index=False)

    return output_file


def plot_excavation_chart(output_dir, labels=None, colors=None, figsize=(4,4)):
    """
    Searches for excavation shapefiles inside output_dir/shape_file_outputs,
    computes areas, plots a donut chart, saves it, and returns DataFrame + image path.
    """

    # === Define expected excavation shapefiles ===
    expected_files = {
        "Planned & active": "planned_and_done_excavation.shp",
        "Unplanned & active": "unplanned_and_done_excavation.shp",
        "Planned & inactive": "planned_and_not_done_excavation.shp"
    }

    shape_dir = os.path.join(output_dir, "shape_file_outputs")
    available_files = {}

    # === Collect available files ===
    for label, fname in expected_files.items():
        fpath = os.path.join(shape_dir, fname)
        if os.path.exists(fpath):
            available_files[label] = fpath

    if not available_files:
        raise FileNotFoundError("No excavation shapefiles found in shape_file_outputs.")

    # === Compute areas (in hectares) ===
    areas = {}
    for label, fpath in available_files.items():
        gdf = gpd.read_file(fpath)
        areas[label] = gdf.geometry.area.sum() / 10000

    # === Prepare chart data ===
    if colors is None:
        colors = ["#0DA21C", "#E71D0E", "#E6D00E"]  

    labels_used = list(areas.keys())
    sizes = list(areas.values())
    chart_colors = colors[:len(sizes)]

    # === Plot donut chart ===
    fig, ax = plt.subplots(figsize=figsize)
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels_used,
        autopct='%1.1f%%',
        startangle=90,
        colors=chart_colors,
        wedgeprops=dict(width=0.4)
    )

    # Reposition autopct text inside donut
    for i, a in enumerate(autotexts):
        ang = (wedges[i].theta2 + wedges[i].theta1) / 2.0
        x = np.cos(np.deg2rad(ang))
        y = np.sin(np.deg2rad(ang))
        a.set_position((0.8*x, 0.8*y))   
        a.set_color("white")
        a.set_weight("bold")

    ax.set(aspect="equal")

    # === Save chart ===
    img_path = os.path.join(output_dir, "excavation_donut_chart.png")
    plt.savefig(img_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

    # === Create DataFrame ===
    df = pd.DataFrame({
        "Category": labels_used,
        "Area_ha": sizes
    })

    return df, img_path


def plot_category_chart(output_dir, category, labels=None, colors=None, figsize=(4,4)):

    # === Define expected shapefiles per category ===
    category_files = {
        "excavation": {
            "Planned & active": "planned_and_done_excavation.shp",
            "Unplanned & active": "unplanned_and_done_excavation.shp",
            "Planned & inactive": "planned_and_not_done_excavation.shp"
        },
        "dump": {
            "Planned & active": "planned_and_done_dump.shp",
            "Unplanned & active": "unplanned_and_done_dump.shp",
            "Planned & inactive": "planned_and_not_done_dump.shp"
        }
    }

    if category not in category_files:
        raise ValueError("Invalid category. Use 'excavation' or 'dump'.")

    expected_files = category_files[category]
    shape_dir = os.path.join(output_dir, "shapefile_outputs")
    available_files = {}

    # === Collect available files ===
    for label, fname in expected_files.items():
        fpath = os.path.join(shape_dir, fname)
        if os.path.exists(fpath):
            available_files[label] = fpath

    if not available_files:
        raise FileNotFoundError(f"No {category} shapefiles found in shapefile_outputs.")

    # === Compute areas (in hectares) ===
    areas = {}
    for label, fpath in available_files.items():
        gdf = gpd.read_file(fpath)
        areas[label] = gdf.geometry.area.sum() / 10000  # m² to ha

    # === Prepare chart data ===
    if colors is None:
        colors = ["#0DA21C", "#E71D0E", "#E6D00E"]

    labels_used = list(areas.keys())
    sizes = list(areas.values())
    chart_colors = colors[:len(sizes)]

    # === Plot donut chart ===
    fig, ax = plt.subplots(figsize=figsize)
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels_used,
        autopct='%1.1f%%',
        startangle=90,
        colors=chart_colors,
        wedgeprops=dict(width=0.4)
    )

    # Reposition autopct text inside donut
    for i, a in enumerate(autotexts):
        ang = (wedges[i].theta2 + wedges[i].theta1) / 2.0
        x = np.cos(np.deg2rad(ang))
        y = np.sin(np.deg2rad(ang))
        a.set_position((0.8 * x, 0.8 * y))
        a.set_color("white")
        a.set_weight("bold")

    ax.set(aspect="equal")
    plt.title(f"{category.capitalize()} Distribution")

    # === Save chart ===
    img_path = os.path.join(output_dir, f"{category}_donut_chart.png")
    plt.savefig(img_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

    # === Create DataFrame ===
    df = pd.DataFrame({
        "Category": labels_used,
        "Area_ha": sizes
    })

    return df, img_path


def update_excel_with_tables_charts(
    input_excel,
    df_excavation,
    img_path_excavation,
    df_dump,
    img_path_dump
):
    # === Load existing workbook ===
    wb = openpyxl.load_workbook(input_excel)
    ws = wb.active

    # Column start positions
    excavation_col = 7
    dump_col = 20

    # === Insert Excavation Table ===
    ws.cell(row=2, column=excavation_col, value="Excavation Summary")
    for r_idx, row in enumerate(dataframe_to_rows(df_excavation, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=excavation_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # === Insert Excavation Chart ===
    img_exc = XLImage(img_path_excavation)
    cell_width = 70   # approx pixel per column
    cell_height = 50  # approx pixel per row
    img_exc.width = 10 * cell_width
    img_exc.height = 10 * cell_height
    ws.add_image(img_exc, ws.cell(row=3 + len(df_excavation) + 2, column=excavation_col).coordinate)

    # === Insert Dump Table ===
    ws.cell(row=2, column=dump_col, value="Dump Summary")
    for r_idx, row in enumerate(dataframe_to_rows(df_dump, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=dump_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # === Insert Dump Chart ===
    img_dump = XLImage(img_path_dump)
    img_dump.width = 10 * cell_width
    img_dump.height = 10 * cell_height
    ws.add_image(img_dump, ws.cell(row=3 + len(df_dump) + 2, column=dump_col).coordinate)

    # === Save updated workbook in same folder as input ===
    input_dir = os.path.dirname(input_excel)
    input_filename = os.path.basename(input_excel)
    name, ext = os.path.splitext(input_filename)

    output_excel = os.path.join(input_dir, f"{name}_updated{ext}")
    wb.save(output_excel)

    return output_excel