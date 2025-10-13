import os
import numpy as np
import rasterio
import pandas as pd
import geopandas as gpd
import shapely
from shapely.ops import split, unary_union
from shapely.geometry import Point, LineString, Polygon, MultiPolygon, MultiLineString
from shapely.ops import unary_union
import matplotlib.pyplot as plt
import ezdxf
from ezdxf.enums import TextEntityAlignment
from ezdxf.math import Vec2

import openpyxl
from openpyxl.styles import PatternFill, Font
import ezdxf
import numpy as np
from ezdxf.math import Vec2
from ezdxf.enums import TextEntityAlignment
import time

## count the section lines
def count_line_features(shapefile_path: str):
    """
    Checks if all features in a shapefile are LineString or MultiLineString.
    If so, returns the count. Otherwise, prints 'Wrong input'.

    Parameters:
        shapefile_path (str): Path to the input shapefile.

    Returns:
        int: Number of LineString/MultiLineString features if valid.
        None: If invalid geometries are found.
    """
    gdf = gpd.read_file(shapefile_path)

    if gdf.geometry.apply(lambda geom: isinstance(geom, (LineString, MultiLineString))).all():
        count = len(gdf)
        print(f"Valid input: {count} LineString/MultiLineString features")
        return None
    else:
        print("Wrong input")
        return None


def clip_single_line_by_boundary(boundary_path:str, line_gdf, output_folder:str):
    # --- Validate inputs ---
    boundary_gdf = gpd.read_file(boundary_path)
    if len(boundary_gdf) != 1:
        print(len(boundary_gdf), "features in boundary_gdf")
        raise ValueError("Boundary GeoDataFrame must contain exactly one feature.")
    if not isinstance(boundary_gdf.geometry.iloc[0], (Polygon, MultiPolygon)):
        raise TypeError("Boundary must be a Polygon or MultiPolygon.")

    if len(line_gdf) != 1:
        raise ValueError("Line GeoDataFrame must contain exactly one feature.")
    if not isinstance(line_gdf.geometry.iloc[0], (LineString, MultiLineString)):
        raise TypeError("Line must be a LineString or MultiLineString.")

    # --- CRS check ---
    if boundary_gdf.crs != line_gdf.crs:
        boundary_gdf = boundary_gdf.to_crs(line_gdf.crs)

    boundary_geom = boundary_gdf.geometry.iloc[0]
    buffered_boundary = boundary_geom.buffer(1e-6)  # Precision buffer
    line_geom = line_gdf.geometry.iloc[0]

    # --- Case 1: Entire line is inside or touching the boundary ---
    if boundary_geom.contains(line_geom) or boundary_geom.touches(line_geom):
        output_geom = line_geom

    # --- Case 2: Line intersects boundary ---
    elif line_geom.intersects(boundary_geom):
        clipped = line_geom.intersection(boundary_geom)

        if isinstance(clipped, LineString):
            start = Point(clipped.coords[0])
            end = Point(clipped.coords[-1])
            if (buffered_boundary.contains(start) or buffered_boundary.touches(start)) and \
               (buffered_boundary.contains(end) or buffered_boundary.touches(end)):
                output_geom = clipped
            else:
                raise ValueError("Clipped line endpoints are not both inside or on boundary.")
        elif isinstance(clipped, MultiLineString):
            found = False
            for part in clipped.geoms:
                start = Point(part.coords[0])
                end = Point(part.coords[-1])
                if (buffered_boundary.contains(start) or buffered_boundary.touches(start)) and \
                   (buffered_boundary.contains(end) or buffered_boundary.touches(end)):
                    output_geom = part
                    found = True
                    break
            if not found:
                raise ValueError("No valid clipped segment lies fully inside or on boundary.")
        else:
            raise TypeError("Unexpected geometry after clipping.")
    else:
        raise ValueError("Line does not intersect or lie within the boundary.")

    # --- Output GeoDataFrame ---
    output_gdf = gpd.GeoDataFrame([line_gdf.iloc[0]], geometry=[output_geom], crs=line_gdf.crs)

    if len(output_gdf) != 1 or not isinstance(output_gdf.geometry.iloc[0], (LineString, MultiLineString)):
        raise ValueError("Output GeoDataFrame must contain exactly one LineString/MultiLineString feature.")

    # --- Save result ---
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "clipped_line.shp")
    output_gdf.to_file(output_path)

    return output_gdf



# ## get the intersecting lines as per boundary polygons
# def intersect_line_with_polygons_return_separately(line_gdf, 
#                                                    output_folder_path: str,
#                                                    section_name: str,
#                                                    planned_and_done_excavation_path: str,
#                                                    unplanned_and_done_excavation_path: str,
#                                                    planned_and_used_dump_path: str,
#                                                    unplanned_and_used_dump_path: str):
#     """
#     Intersects a single line with 4 polygon shapefiles.
#     Saves intersected line shapefiles inside a section-named subfolder or default 'intersecting_lines'.
#     If no intersection, creates an empty shapefile with correct schema and CRS.
    
#     Returns:
#         tuple of 4 GeoDataFrames (may be empty but not None)
#     """

#     os.makedirs(output_folder_path, exist_ok=True)
    
#     def process(name, poly_path):
#         poly_gdf = gpd.read_file(poly_path)
#         if poly_gdf.empty:
#             print(f"{name}: Polygon file is empty.")
#             return gpd.GeoDataFrame(geometry=[], crs=line_gdf.crs)

#         # Reproject to match CRS
#         if poly_gdf.crs != line_gdf.crs:
#             poly_gdf = poly_gdf.to_crs(line_gdf.crs)

#         poly_union = poly_gdf.unary_union
#         intersection = line_geom.intersection(poly_union)

#         segments = []
#         if intersection.is_empty:
#             print(f"{name}: No intersection.")
#         elif isinstance(intersection, LineString):
#             segments.append(intersection)
#         elif isinstance(intersection, MultiLineString):
#             segments.extend(intersection.geoms)
#         else:
#             print(f"{name}: Unsupported geometry type {type(intersection)}")

#         # Prepare GeoDataFrame
#         intersect_gdf = gpd.GeoDataFrame(geometry=segments, crs=poly_gdf.crs)

#         if not intersect_gdf.empty:
#             # Assign temporary ID for spatial join
#             intersect_gdf["temp_id"] = intersect_gdf.index
#             intersect_gdf = gpd.sjoin(intersect_gdf, poly_gdf[["geometry", "Section_Na"]], how="left", predicate="intersects")
#             intersect_gdf = intersect_gdf.drop(columns=["index_right", "temp_id"])

#         # Write to file (even if empty)
#         output_path = os.path.join(output_folder_path, f"section_{section_name}_{name}.shp")
#         intersect_gdf.to_file(output_path)
#         #print(f"{name}: Exported {len(intersect_gdf)} segments to {output_path}")

#         return intersect_gdf


#     # Ensure only one line feature is present
#     if len(line_gdf) != 1:
#         print(len(line_gdf))
#         raise ValueError("Line GeoDataFrame must contain exactly one feature.")
#     line_geom = line_gdf.geometry.iloc[0]

#     # Process all polygon layers
#     lines_planned_and_done_excavation = process("planned_and_done_excavation", planned_and_done_excavation_path)
#     lines_unplanned_and_done_excavation = process("unplanned_and_done_excavation", unplanned_and_done_excavation_path)
#     lines_planned_and_used_dump = process("planned_and_used_dump", planned_and_used_dump_path)
#     lines_unplanned_and_used_dump = process("unplanned_and_used_dump", unplanned_and_used_dump_path)

#     return (
#         lines_planned_and_done_excavation,
#         lines_unplanned_and_done_excavation,
#         lines_planned_and_used_dump,
#         lines_unplanned_and_used_dump
#     ), section_name




############### updated function ##########
def intersect_line_with_polygons_return_separately(line_gdf, 
                                                   output_folder_path: str,
                                                   section_name: str,
                                                   planned_and_done_excavation_path: str = None,
                                                   unplanned_and_done_excavation_path: str = None,
                                                   planned_and_used_dump_path: str = None,
                                                   unplanned_and_used_dump_path: str = None):
    """
    Intersects a single line with up to 4 polygon shapefiles (if provided).
    Saves intersected line shapefiles inside a section-named subfolder or default 'intersecting_lines'.
    If no intersection or file is None, creates an empty shapefile with correct schema and CRS.
    
    Returns:
        tuple of 4 GeoDataFrames (may be empty but not None), and section_name
    """
    os.makedirs(output_folder_path, exist_ok=True)

    def process(name, poly_path):
        output_path = os.path.join(output_folder_path, f"section_{section_name}_{name}.shp")

        # Return empty GeoDataFrame if path is None or file does not exist
        if poly_path is None or not os.path.exists(poly_path):
            print(f"{name}: File path is None or does not exist.")
            empty_gdf = gpd.GeoDataFrame(geometry=[], crs=line_gdf.crs)
            empty_gdf.to_file(output_path)
            return empty_gdf

        poly_gdf = gpd.read_file(poly_path)
        if poly_gdf.empty:
            print(f"{name}: Polygon file is empty.")
            empty_gdf = gpd.GeoDataFrame(geometry=[], crs=line_gdf.crs)
            empty_gdf.to_file(output_path)
            return empty_gdf

        # Reproject to match CRS
        if poly_gdf.crs != line_gdf.crs:
            poly_gdf = poly_gdf.to_crs(line_gdf.crs)

        poly_union = poly_gdf.unary_union
        intersection = line_geom.intersection(poly_union)

        segments = []
        if intersection.is_empty:
            print(f"{name}: No intersection.")
        elif isinstance(intersection, LineString):
            segments.append(intersection)
        elif isinstance(intersection, MultiLineString):
            segments.extend(intersection.geoms)
        else:
            print(f"{name}: Unsupported geometry type {type(intersection)}")

        # Prepare GeoDataFrame
        intersect_gdf = gpd.GeoDataFrame(geometry=segments, crs=poly_gdf.crs)

        if not intersect_gdf.empty:
            intersect_gdf["temp_id"] = intersect_gdf.index
            intersect_gdf = gpd.sjoin(intersect_gdf, poly_gdf[["geometry", "Section_Na"]], how="left", predicate="intersects")
            intersect_gdf = intersect_gdf.drop(columns=["index_right", "temp_id"])

        # Save output
        intersect_gdf.to_file(output_path)
        return intersect_gdf

    # Ensure only one line feature is present
    if len(line_gdf) != 1:
        raise ValueError("Line GeoDataFrame must contain exactly one feature.")
    line_geom = line_gdf.geometry.iloc[0]

    # Process each layer
    lines_planned_and_done_excavation = process("planned_and_done_excavation", planned_and_done_excavation_path)
    lines_unplanned_and_done_excavation = process("unplanned_and_done_excavation", unplanned_and_done_excavation_path)
    lines_planned_and_used_dump = process("planned_and_used_dump", planned_and_used_dump_path)
    lines_unplanned_and_used_dump = process("unplanned_and_used_dump", unplanned_and_used_dump_path)

    return (
        lines_planned_and_done_excavation,
        lines_unplanned_and_done_excavation,
        lines_planned_and_used_dump,
        lines_unplanned_and_used_dump
    ), section_name



# ---------- Step 1: Elevation querying helper ----------
def elevation_at_point(dtm_src, x, y):
    """Query elevation value from DTM raster at (x, y) location."""
    q_ele = list(dtm_src.sample([(x, y)]))[0][0]
    return q_ele

# ---------- Step 2: Elevation sampling function ----------
def get_elevation_points_from_lines(dtm_path: str, clipped_gdf, interval: float = 0.01):
    """
    For each point interpolated along lines, fetch elevation from DTM.
    Returns: List of (x, y, z) tuples.
    """
    lines_gdf = clipped_gdf

    with rasterio.open(dtm_path) as dtm_src:
        if lines_gdf.crs != dtm_src.crs:
            print("Reprojecting line geometries to match DTM CRS...")
            lines_gdf = lines_gdf.to_crs(dtm_src.crs)

        points_with_elevation = []

        for geom in lines_gdf.geometry:
            if isinstance(geom, LineString):
                length = geom.length
                num_points = int(length // interval) + 1
                distances = [i * interval for i in range(num_points + 1)]

                for d in distances:
                    pt = geom.interpolate(d)
                    x, y = pt.x, pt.y
                    z = elevation_at_point(dtm_src, x, y)
                    points_with_elevation.append((x, y, z))

        return points_with_elevation
    

def extract_segment_df(intersect_line, clipped_line_gdf, dtm1_path, dtm2_path):
    gdf = intersect_line
    if gdf.empty:
        return pd.DataFrame()  # Return empty if no data

    if gdf.crs != clipped_line_gdf.crs:
        gdf = gdf.to_crs(clipped_line_gdf.crs)

    main_line = clipped_line_gdf.geometry.iloc[0]
    segment_points = []

    with rasterio.open(dtm1_path) as dtm1, rasterio.open(dtm2_path) as dtm2:
        for idx, seg in enumerate(gdf.geometry):
            area_name = gdf.iloc[idx].get("Section_Na", None)
            for pt in [seg.interpolate(0), seg.interpolate(seg.length)]:
                chainage = main_line.project(pt)
                elev1 = elevation_at_point(dtm1, pt.x, pt.y)
                elev2 = elevation_at_point(dtm2, pt.x, pt.y)
                segment_points.append({
                    "chainage": chainage,
                    "elevation_itr1": elev1,
                    "elevation_itr2": elev2,
                    "Section_Na": area_name
                })
    #print(segment_points)
    return pd.DataFrame(segment_points)


def plot_multiple_intersections_on_elevation(df1, dtm_itr1_year_ip, df2, dtm_itr2_year_ip, segment_data_dict, section_name, output_folder_path):
    import matplotlib.pyplot as plt
    import pandas as pd
    import os

    start_label, end_label = section_name.split("_")
    fig, ax = plt.subplots(figsize=(20, 6))
    fig.patch.set_facecolor('white')  # Figure background
    ax.set_facecolor('white')         # Plot (axes) background

    # Plot base elevation profiles
    ax.plot(df1["chainage"], df1["z"], label=f"Elevation ITR {dtm_itr1_year_ip}", color="blue", linewidth=2)
    ax.plot(df2["chainage"], df2["z"], label=f"Elevation ITR {dtm_itr2_year_ip}", color="orange", linewidth=2)

    intersection_chainages = []

    for label, (segment_df, _) in segment_data_dict.items():
        if segment_df is None or segment_df.empty:
            continue

        # Plot the points
        ax.scatter(segment_df["chainage"], segment_df["elevation_itr1"], color="black", s=5, zorder=5)
        ax.scatter(segment_df["chainage"], segment_df["elevation_itr2"], color="black", s=5, zorder=5)

        for idx, row in segment_df.iterrows():
            area_name = str(row.get("Section_Na", ""))
            ch = row["chainage"]
            z = row["elevation_itr1"]

            # Decide fill color (keep original logic)
            if area_name.startswith("PEA") or area_name.startswith("PDA"):
                fill_color = "#9ACD32"  # planned
                text_color = "#008000"  # dark green
            elif area_name.startswith("UEA") or area_name.startswith("UDA"):
                fill_color = "#fd5c63"  # unplanned
                text_color = "#FF0000"  # dark red
            else:
                fill_color = "gray"
                text_color = "black"

            # Annotation above the point
            if pd.notnull(area_name):
                ha = 'left' if idx % 2 == 0 else 'right'
                offset_x = 1.2 if idx % 2 == 0 else -1

                ax.annotate(area_name, xy=(ch, z), xytext=(offset_x, 18), textcoords='offset points',
                            fontsize=6, color=text_color, ha=ha, va='bottom', rotation=90)

            # Collect chainage for vertical line
            intersection_chainages.append(ch)

        # Fill area between elevation lines with original logic
        for i in range(0, len(segment_df), 2):
            if i + 1 >= len(segment_df):
                continue

            ch_start = min(segment_df.iloc[i]["chainage"], segment_df.iloc[i+1]["chainage"])
            ch_end   = max(segment_df.iloc[i]["chainage"], segment_df.iloc[i+1]["chainage"])

            mask = (df1["chainage"] >= ch_start) & (df1["chainage"] <= ch_end)
            x = df1.loc[mask, "chainage"]
            y1 = df1.loc[mask, "z"]
            y2 = df2.loc[mask, "z"]

            ax.fill_between(x, y1, y2, color=fill_color, alpha=0.4, label=label if i == 0 else None)

    # Draw vertical dashed lines in cyan only
    for ch in intersection_chainages:
        ax.axvline(x=ch, color='gray', linestyle='--', linewidth=1)

    # Add section labels
    ax.annotate(start_label, xy=(df1["chainage"].iloc[0], df1["z"].iloc[0]),
                textcoords="offset points", xytext=(-20, 10), fontsize=10, color="Black", weight='bold')
    ax.annotate(end_label, xy=(df1["chainage"].iloc[-1], df1["z"].iloc[-1]),
                textcoords="offset points", xytext=(10, 10), fontsize=10, color="Black", weight='bold')

    ax.set_xlabel("Chainage (m)")
    ax.set_ylabel("Elevation (m)")
    new_section_name = section_name.replace("_", "")
    ax.set_title(f"Elevation vs. Chainage – Section {new_section_name}")
    ax.grid(False)
    ax.legend()
    plt.tight_layout()

    # Save plot
    section_folder_name = f"section_{section_name}_plot"
    output_path = os.path.join(output_folder_path, section_folder_name)
    os.makedirs(output_path, exist_ok=True)
    file_path = os.path.join(output_path, f"{section_folder_name}_plot.png")
    plt.savefig(file_path, dpi=300, facecolor='white')
    plt.close()
    print(f"Plot saved to: {file_path}")
    
    return None



def compute_maximum_deviation_by_intersection(segment_data_dict):
    max_deviations = {}
    for label, (segment_df, _) in segment_data_dict.items():
        #print(f"\n Checking: {label}")

        if segment_df is None or segment_df.empty:
            #print(f" Empty or None dataframe for {label}")
            max_deviations[label] = None
            continue

        if "elevation_itr1" not in segment_df.columns or "elevation_itr2" not in segment_df.columns:
            #print(f" Missing elevation columns in {label}")
            max_deviations[label] = None
            continue

        segment_df["deviation"] = segment_df["elevation_itr1"] - segment_df["elevation_itr2"]

        if segment_df["deviation"].abs().dropna().empty:
            #print(f" All deviation values are NaN for {label}")
            max_deviations[label] = None
            continue

        max_dev_idx = segment_df["deviation"].abs().idxmax()
        max_deviation = segment_df.loc[max_dev_idx, "deviation"]
        #print(f" Max deviation in {label}: {max_deviation}")
        max_deviations[label] = round(max_deviation, 3)

    return max_deviations

def export_all_section_deviation_summary_csv(section_results, output_folder_path, csv_filename="elevation_deviation_summary.csv"):
    summary_rows = []

    for idx, (section_name, segment_data_dict) in enumerate(section_results, start=1):
        deviations = compute_maximum_deviation_by_intersection(segment_data_dict)

        # Extract unplanned deviations
        unplanned_excav = deviations.get("unplanned_and_done_excavation")
        unplanned_dump = deviations.get("unplanned_and_used_dump")

        # Determine status and max deviation
        if unplanned_excav is None and unplanned_dump is None:
            status = "No Deviation detected"
            max_dev = "NA"
        elif unplanned_excav is None and unplanned_dump is not None:
            status = "Deviation detected in Dump"
            max_dev = abs(unplanned_dump)
        elif unplanned_excav is not None and unplanned_dump is None:
            status = "Deviation detected in Excavation"
            max_dev = abs(unplanned_excav)
        else:
            status = "Deviation detected in Both"
            max_dev = max(abs(unplanned_excav), abs(unplanned_dump))
        
        print(f"Section: {section_name}, Status: {status}, Max Deviation: {max_dev}")
        print(type(max_dev), "max_dev type")
        new_section_name = section_name.replace("_", "")


        # Construct row
        row = {
            "Sr.No.": idx,
            "Section Name": new_section_name,
            #"Planned and Excavation": round(deviations.get("planned_and_done_excavation", 0), 2) if deviations.get("planned_and_done_excavation") is not None else "NA",
            #"Unplanned and Excavation": round(unplanned_excav, 2) if unplanned_excav is not None else "NA",
            #"Planned and Used Dump": round(deviations.get("planned_and_used_dump", 0), 2) if deviations.get("planned_and_used_dump") is not None else "NA",
            #"Unplanned and Used Dump": round(unplanned_dump, 2) if unplanned_dump is not None else "NA",
            "Status": status,
            "Max Deviation": round(float(max_dev), 2) if isinstance(max_dev, (int, float, np.floating)) else max_dev
        }

        summary_rows.append(row)

    df_summary = pd.DataFrame(summary_rows)
    os.makedirs(output_folder_path, exist_ok=True)

    # Save to CSV
    csv_path = os.path.join(output_folder_path, csv_filename)
    df_summary.to_csv(csv_path, index=False)
    print(f"\n Summary CSV saved: {csv_path}")

    # Save to Excel with formatting
    excel_path = os.path.join(output_folder_path, csv_filename.replace(".csv", ".xlsx"))
    df_summary.to_excel(excel_path, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Fill styles
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    dark_red_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # Bold font
    bold_font = Font(bold=True)

    # Color and bold the header row
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = bold_font

    # Apply formatting to the Status column
    for row in range(2, ws.max_row + 1):
        status_cell = ws[f"C{row}"]
        status_cell.font = bold_font

        if status_cell.value == "Deviation detected in Both":
            status_cell.fill = dark_red_fill
        elif status_cell.value in ["Deviation detected in Dump", "Deviation detected in Excavation"]:
            status_cell.fill = red_fill
        elif status_cell.value == "No Deviation detected":
            status_cell.fill = green_fill

    wb.save(excel_path)
    print(f" Excel with formatting saved: {excel_path}")


# ### Export elevation profiles to DXF with fill areas between ITRs
# def export_elevation_profiles_to_dxf_with_fill(df1, df2, segment_data_dict, section_name, output_path="elevation_profiles.dxf"):
#     doc = ezdxf.new(dxfversion="R2010")
#     msp = doc.modelspace()

#     start_label, end_label = section_name.split("_")

#     # Define color map for segment labels (to match matplotlib fill color if you have any specific)
#     color_lookup = {
#         "Zone 1": 1,  # red
#         "Zone 2": 2,  # yellow
#         "Zone 3": 3,  # green
#         "Zone 4": 4,  # cyan
#         "Zone 5": 5,  # blue
#         "Zone 6": 6,  # magenta
#         # Add more if needed or dynamically assign
#     }

#     # Plot ITR1
#     points1 = list(zip(df1["chainage"], df1["z"]))
#     msp.add_lwpolyline(points1, dxfattribs={"color": 5})  # blue
#     msp.add_text("ITR 1", dxfattribs={"height": 5, "color": 5}).set_placement(
#         (df1["chainage"].iloc[-1], df1["z"].iloc[-1]), align=TextEntityAlignment.LEFT
#     )

#     # Plot ITR2
#     points2 = list(zip(df2["chainage"], df2["z"]))
#     msp.add_lwpolyline(points2, dxfattribs={"color": 3})  # green
#     msp.add_text("ITR 2", dxfattribs={"height": 5, "color": 3}).set_placement(
#         (df2["chainage"].iloc[-1], df2["z"].iloc[-1]), align=TextEntityAlignment.LEFT
#     )

#     # Draw vertical dashed lines at intersection points
#     for label, (segment_df, _) in segment_data_dict.items():
#         if segment_df is None or segment_df.empty:
#             continue
#         for ch in segment_df["chainage"]:
#             msp.add_line((ch, df1["z"].min() - 5), (ch, df1["z"].max() + 5),
#                          dxfattribs={"linetype": "DASHED", "color": 4})

#     # Fill zones between ITR1 and ITR2 based on segment_data_dict
#     for label, (segment_df, color_code) in segment_data_dict.items():
#         if segment_df is None or segment_df.empty:
#             continue
#         print(color_code, "color_code")
#         # assign final dxf color based on label
#         if color_code in ["green"]:
#             dxf_color = 3 ## green
#         elif color_code in ["red"]:
#             dxf_color = 1 ## red
#         else: 
#             dxf_color = 2 ## yellow


#         for i in range(0, len(segment_df), 2):
#             if i + 1 >= len(segment_df):
#                 continue

#             ch_start = min(segment_df.iloc[i]["chainage"], segment_df.iloc[i + 1]["chainage"])
#             ch_end = max(segment_df.iloc[i]["chainage"], segment_df.iloc[i + 1]["chainage"])

#             # Filter chainage range from df1 and df2
#             mask = (df1["chainage"] >= ch_start) & (df1["chainage"] <= ch_end)
#             x = df1.loc[mask, "chainage"].values
#             y1 = df1.loc[mask, "z"].values
#             y2 = df2.loc[mask, "z"].values

#             if len(x) < 2:
#                 continue

#             polygon_points = [Vec2(c, z) for c, z in zip(x, y1)] + [Vec2(c, z) for c, z in zip(x[::-1], y2[::-1])]
#             hatch = msp.add_hatch(dxfattribs={"color": dxf_color})
#             hatch.paths.add_polyline_path(polygon_points, is_closed=True)

#             ## Set true_color if RGB is provided, else fallback to indexed color
#             #if isinstance(color, int) and color > 256:
#             #    hatch.dxf.true_color = color  # RGB format like 0x90EE90
#             #else:
#             #    hatch.dxf.color = color_lookup.get(label, 3)  # fallback

#     # X and Y axis
#     x_min, x_max = df1["chainage"].min(), df1["chainage"].max()
#     y_min = min(df1["z"].min(), df2["z"].min()) - 5
#     y_max = max(df1["z"].max(), df2["z"].max()) + 5

#     # X-axis
#     msp.add_line((x_min, y_min), (x_max + 10, y_min), dxfattribs={"color": 7})
#     msp.add_text("Chainage (m)", dxfattribs={"height": 10, "color": 7}).set_placement(
#         (x_max + 12, y_min), align=TextEntityAlignment.LEFT
#     )

#     # Y-axis
#     msp.add_line((x_min, y_min - 5), (x_min, y_max), dxfattribs={"color": 7})
#     msp.add_text("Elevation (m)", dxfattribs={"height": 10, "color": 7}).set_placement(
#         (x_min, y_max + 2), align=TextEntityAlignment.LEFT
#     )

#     # Add tick values on X and Y axes
#     for ch_tick in np.linspace(x_min, x_max, num=10):
#         msp.add_text(f"{int(ch_tick)}", dxfattribs={"height": 5, "color": 7}).set_placement(
#             (ch_tick, y_min - 2.5), align=TextEntityAlignment.MIDDLE
#         )

#     for z_tick in np.linspace(y_min + 5, y_max, num=8):
#         msp.add_text(f"{int(z_tick)}", dxfattribs={"height": 5, "color": 7}).set_placement(
#             (x_min - 5, z_tick), align=TextEntityAlignment.MIDDLE_RIGHT
#         )

#     # Add section start and end labels
#     msp.add_text(start_label, dxfattribs={"height": 5, "color": 5}).set_placement(
#         (df1["chainage"].iloc[0] - 5, df1["z"].iloc[0] + 5), align=TextEntityAlignment.LEFT
#     )
#     msp.add_text(end_label, dxfattribs={"height": 5, "color": 5}).set_placement(
#         (df1["chainage"].iloc[-1] + 5, df1["z"].iloc[-1] + 5), align=TextEntityAlignment.LEFT
#     )


#     # Add title to the plot
#     title_text = f"Section: {section_name} Elevation Plot"
#     title_x = (x_min + x_max) / 2
#     title_y = y_max + 10  # a bit above the top of the plot

#     msp.add_text(title_text, dxfattribs={"height": 15, "color": 5}).set_placement(
#         (title_x, title_y), align=TextEntityAlignment.MIDDLE_CENTER
#     )

#     # Save DXF
#     doc.saveas(output_path)
#     print(f"DXF saved with fills: {output_path}")


def export_to_dxf_with_colors_updated(df1, df2, segment_data_dict, output_dxf_path, section_name):
    
    t_1 = time.perf_counter()

    code_mapping = {
    "planned_and_done_excavation": ("Excavation", "Planned"),
    "unplanned_and_done_excavation": ("Excavation", "Unplanned"),
    "planned_and_used_dump": ("Dump", "Planned"),
    "unplanned_and_used_dump": ("Dump", "Unplanned"),
}

    def get_zone_info_from_code(code):
        return code_mapping.get(code, ("Unknown", "Unknown"))

    doc = ezdxf.new(dxfversion='R2018')
    msp = doc.modelspace()

    # Plot ITR1 (blue)
    #points1 = list(zip(df1["chainage"], df1["z"]))
    t0 = time.perf_counter()
    points1 = np.column_stack((df1["chainage"].values, df1["z"].values))
    msp.add_lwpolyline(points1, dxfattribs={"color": 5})  # Blue
    t1 = time.perf_counter()
    print(f"Plotted ITR 1 in {t1 - t0:.3f} seconds")

    # coordinates for ITR! Label
    itr1_x = df1["chainage"].iloc[-1]
    itr1_y = df1["z"].iloc[-1]

    msp.add_text("ITR 1", dxfattribs={"height": 5, "color": 5}).set_placement(
        (itr1_x,itr1_y + 2.5), align=TextEntityAlignment.LEFT
    )
    print("Placed ITR 1 label at position:", itr1_x, itr1_y)
    print()

    # Plot ITR2 (orange - true color)
    #points2 = list(zip(df2["chainage"], df2["z"]))
    t0 = time.perf_counter()
    points2 = np.column_stack((df2["chainage"].values, df2["z"].values))
    poly2 = msp.add_lwpolyline(points2)
    t1 = time.perf_counter()
    print(f"Plotted ITR 2 in {t1 - t0:.3f} seconds")
    poly2.dxf.true_color = 0xFFA500  # Orange

    itr2_x = itr1_x
    itr2_y = itr1_y - 5

    text2 = msp.add_text("ITR 2", dxfattribs={"height": 5})
    text2.set_placement((itr2_x,itr2_y), align=TextEntityAlignment.LEFT)
    text2.dxf.true_color = 0xFFA500
    print("Placed ITR 2 label at position:", itr2_x, itr2_y)
    print()

    if pd.notna(df2["chainage"].iloc[-1]) and pd.notna(df2["z"].iloc[-1]):
        text2 = msp.add_text("ITR 2", dxfattribs={"height": 5})
        text2.set_placement((df2["chainage"].iloc[-1], df2["z"].iloc[-1]), align=TextEntityAlignment.LEFT)
        text2.dxf.true_color = 0xFFA500
        print("Placed ITR 2 label at position:", df2["chainage"].iloc[-1], df2["z"].iloc[-1])
    else:
        print("Cannot place ITR 2 label: NaN in position")
    print()

    min_z = min(df1["z"].min(), df2["z"].min())
    max_z = max(df1["z"].max(), df2["z"].max())

    for label, (segment_df, _) in segment_data_dict.items():
        print("Label Passed:",label)
        zone, status = get_zone_info_from_code(label)
        print("Zone:", zone, "Status:", status)
        if segment_df is None or segment_df.empty:
            continue

        #zone, status = get_zone_info_from_code(label)

        # Define RGB color for hatching
        if status == "Planned":
            fill_rgb = 0x00FF00  # Green
        elif status == "Unplanned":
            fill_rgb = 0xFF0000  # Red
        else:
            fill_rgb = 0xAAAAAA  # Gray

        for ch in segment_df["chainage"]:
            msp.add_line(
                (ch, min_z - 5), (ch, max_z + 5),
                dxfattribs={"linetype": "DASHED", "color": 8}
            )

        for i in range(0, len(segment_df), 2):
            if i + 1 >= len(segment_df):
                continue
            ch1 = segment_df.iloc[i]["chainage"]
            ch2 = segment_df.iloc[i + 1]["chainage"]
            ch_start, ch_end = min(ch1, ch2), max(ch1, ch2)

            mask = (df1["chainage"] >= ch_start) & (df1["chainage"] <= ch_end)
            x = df1.loc[mask, "chainage"].values
            y1 = df1.loc[mask, "z"].values
            y2 = df2.loc[mask, "z"].values

            if len(x) < 2:
                continue

            polygon_points = [Vec2(c, z) for c, z in zip(x, y1)] + [Vec2(c, z) for c, z in zip(x[::-1], y2[::-1])]
            hatch = msp.add_hatch()
            hatch.paths.add_polyline_path(polygon_points, is_closed=True)
            hatch.dxf.true_color = fill_rgb  # <- Fixed here

        for idx, row in segment_df.iterrows():
            ch = row["chainage"]
            z = row["elevation_itr1"]
            name = str(row.get("Section_Na", label))
            offset_x = 1.2 if idx % 2 == 0 else -1
            align = TextEntityAlignment.RIGHT if idx % 2 != 0 else TextEntityAlignment.LEFT

            msp.add_circle(center=(ch, z), radius=0.3, dxfattribs={"color": 0})

            label_text = msp.add_text(
                name,
                dxfattribs={"height": 2.5, "rotation": 90}
            )
            label_text.dxf.true_color = fill_rgb
            label_text.set_placement((ch + offset_x, z + 1), align=align)

        if not segment_df.empty:
            mid_idx = len(segment_df) // 2
            mid_chainage = segment_df.iloc[mid_idx]["chainage"]
            mid_elev = (segment_df.iloc[mid_idx]["elevation_itr1"] + segment_df.iloc[mid_idx]["elevation_itr2"]) / 2

            label = msp.add_text(f"{status} {zone}", dxfattribs={"height": 5})
            label.dxf.true_color = fill_rgb
            label.set_placement((mid_chainage, mid_elev), align=TextEntityAlignment.MIDDLE_CENTER)

    # Start/End Labels
    start_label, end_label = section_name.split("_")
    msp.add_text(start_label, dxfattribs={"height": 5, "color": 0}).set_placement(
        (df1["chainage"].iloc[0], df1["z"].iloc[0] + 5), align=TextEntityAlignment.LEFT
    )
    msp.add_text(end_label, dxfattribs={"height": 5, "color": 0}).set_placement(
        (df1["chainage"].iloc[-1], df1["z"].iloc[-1] + 5), align=TextEntityAlignment.LEFT
    )

    # Title
    msp.add_text(
        f"Elevation vs Chainage – Section {section_name.replace('_','')}",
        dxfattribs={"height": 6, "color": 0}
    ).set_placement(
        (df1["chainage"].mean(), max_z + 10), align=TextEntityAlignment.MIDDLE_CENTER
    )

    # Axes
    x_min, x_max = df1["chainage"].min(), df1["chainage"].max()
    y_min, y_max = min_z - 5, max_z + 5

    msp.add_line((x_min, y_min), (x_max, y_min), dxfattribs={"color": 0})
    msp.add_line((x_min, y_min), (x_min, y_max), dxfattribs={"color": 0})

    x_ticks = np.arange(np.floor(x_min / 100) * 100, np.ceil(x_max / 100) * 100 + 1, 100)
    for xt in x_ticks:
        msp.add_line((xt, y_min), (xt, y_min - 1), dxfattribs={"color": 8})
        msp.add_text(f"{int(xt)}", dxfattribs={"height": 2.5, "color": 8}).set_placement(
            (xt, y_min - 3), align=TextEntityAlignment.MIDDLE_CENTER
        )

    y_ticks = np.arange(np.floor(y_min / 10) * 10, np.ceil(y_max / 10) * 10 + 1, 10)
    for yt in y_ticks:
        msp.add_line((x_min, yt), (x_min - 1, yt), dxfattribs={"color": 8})
        msp.add_text(f"{int(yt)}", dxfattribs={"height": 2.5, "color": 8}).set_placement(
            (x_min - 3, yt), align=TextEntityAlignment.MIDDLE_CENTER
        )

    msp.add_text("Chainage (m)", dxfattribs={"height": 3.5, "color": 0}).set_placement(
        ((x_min + x_max) / 2, y_min - 6), align=TextEntityAlignment.MIDDLE_CENTER
    )
    msp.add_text("Elevation (m)", dxfattribs={"height": 3.5, "color": 0, "rotation": 90}).set_placement(
        (x_min - 8, (y_min + y_max) / 2), align=TextEntityAlignment.MIDDLE_CENTER
    )

    doc.saveas(output_dxf_path)
    print(f"DXF saved at: {output_dxf_path}")

    t_2 = time.perf_counter()
    print(f"Total time taken to export DXF: {t_2 - t_1:.3f} seconds")
    return None
